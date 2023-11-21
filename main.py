# main.py

from kivymd.app import MDApp
from kivy.lang import Builder
from kivy.core.window import Window
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDFlatButton
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivymd.uix.datatables import MDDataTable
from kivy.metrics import dp
from kivymd.uix.button import MDRaisedButton
import openpyxl
from openpyxl import Workbook
import pathlib

Window.size = (450, 720)


class LoginApp(MDApp):
    def build(self):
        self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette = "BlueGray"
        return Builder.load_string('''
Screen:
    MDCard:
        size_hint : None,None
        size : 350,600
        pos_hint : {'center_x' : 0.5, 'center_y' : 0.5}
        padding : 25
        spacing : 25
        orientation : 'vertical'

        MDLabel:
            id:Login_label
            text:"Login"
            font_size: 25
            halign : 'center'
            size_hint_y : None
            height: self.texture_size[1]
            padding_y: 10
            pos_hint : {'center_x' : 0.5, 'center_y' : 0.5}

        MDTextField:
            id:username
            hint_text:"Username"
            icon_right:"account"
            size_hint_x: None
            width: 250
            font_size: 18
            pos_hint: {"center_x": 0.5}

        MDTextField:
            id:password
            hint_text:"Password"
            icon_right:"eye-off"
            size_hint_x: None
            width: 250
            font_size: 18
            pos_hint: {"center_x": 0.5}
            password:True

        Widget:
            size_hint_y: None
            height:5

        MDRoundFlatButton:
            text:"Login"
            font_size: 18
            pos_hint : {'center_x' : 0.5, 'center_y' : 0.5}
            width: 250
            halign :'center'
            height : 15
            size_hint_x : 0.9
            on_press: app.login()

        MDRoundFlatButton:
            text:"Reset"
            halign :'center'
            font_size: 18
            height:15
            pos_hint : {'center_x' : 0.5,  'center_y' : 0.5}
            width: 250
            size_hint_x : 0.9
            on_press: app.reset()

        Widget:
            size_hint_y: None
            height:9

        Widget:
            size_hint_y: None
            height:9

        Widget:
            size_hint_y: None
            height:9
''')

    def login(self):
        user = self.root.ids.username.text
        passcode = self.root.ids.password.text
        close_button = MDFlatButton(text="Close", on_press=self.close_dialog2)
        dialog = MDDialog(title="Invalid Username", size_hint=(1, 1), text="Enter a valid entry")
        dialog2 = MDDialog(title="Login Successful", size_hint=(1, 1), text="Logging in")
        if user == 'Samvit' and passcode == 'yash1234*':
            dialog2.buttons = [close_button]
            dialog2.open()
            self.open_main_file()
        elif user == '' or passcode == '':
            dialog.open()
        else:
            dialog.open()

    def close_dialog2(self, *args):
        self.dialog2.dismiss()

    def open_main_file(self):
        self.root.clear_widgets()
        self.stop()
        data_app = DataEntryApp()
        data_app.run()

    def reset(self):
        self.root.ids.username.text=""
        self.root.ids.password.text=""


class DataEntryApp(MDApp):
    def build(self):
        self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette = "DeepPurple"
        return Builder.load_string('''
Screen:
    MDBoxLayout:
        orientation: 'vertical'
        padding: 16
        spacing: 10

        MDBoxLayout:
            padding: 20
            spacing: 10
            pos_hint: {'center_y': 0.5}
            size_hint_y: 0.3

            MDTextField:
                hint_text: "Name"
                mode: "rectangle"
                id: name
                icon_right: "account-outline"
                width: 250
                font_size: 18
                pos_hint: {"center_x": 0.5}

            MDTextField:
                hint_text: "Standard"
                mode: "rectangle"
                id: standard
                icon_right: "list-box-outline"
                width: 250
                font_size: 18

            MDRaisedButton:
                text: "Add"
                md_bg_color: "#333333"
                size_hint_y: 1.2
                on_press: app.add()

        MDCard:
            size_hint: None, None
            size: 375, 500
            pos_hint: {'center_x': 0.5, 'center_y': 0.5}
            padding: 15
            spacing: 25

            MDScrollView:
                id: card_layout
                do_scroll_x: False

        MDBoxLayout:
            padding: 20
            spacing: 10
            pos_hint: {'center_y': 0.5}
            size_hint_y: 0.3
            MDRaisedButton:
                text: "Search"
                icon_right: "database-search-outline"
                size_hint_y: 1.2
                size_hint_x: 2
                md_bg_color: "#333333"
                pos_hint: {'center_y': 0.5}
                on_press: app.search()
            MDRaisedButton:
                id: reset_button
                text: "Reset Table"
                icon_right: "refresh"
                size_hint_y: 1.2
                size_hint_x: 2
                md_bg_color: "#333333"
                pos_hint: {'center_y': 0.5}
                on_press: app.reset_table()
''')

    def add(self):
        name = self.root.ids.name.text
        standard = self.root.ids.standard.text

        if name == '' or standard == '':
            print("Fill all fields")
        else:
            print(name)
            print(standard)

            file = openpyxl.load_workbook('Backend_data.xlsx')
            sheet = file.active
            sheet.cell(column=1, row=sheet.max_row + 1, value=name)
            sheet.cell(column=2, row=sheet.max_row, value=standard)

            file.save(r'Backend_data.xlsx')
            self.reset_table()

            self.root.ids.name.text = ''
            self.root.ids.standard.text = ''

    def reset_search_buttons(self, search_button_visible=True):
        self.root.ids.search_button.opacity = 1 if search_button_visible else 0
        self.root.ids.reset_button.opacity = 1 if not search_button_visible else 0

    def search(self):
        name = self.root.ids.name.text
        standard = self.root.ids.standard.text

        if name == '' and standard == '':
            print("Enter at least one value for search")
        else:
            data = self.read_excel_data('Backend_data.xlsx')

            if name:
                data = [row for row in data if name.lower() in row[0].lower()]
            if standard:
                data = [row for row in data if standard.lower() in str(row[1]).lower()]

            cols = ["Name", "Standard"]
            values = data

            table = MDDataTable(
                pos_hint={'center_x': 0.2, 'center_y': 0.2},
                column_data=[(col, dp(35)) for col in cols],
                row_data=values,
                use_pagination=True,
                padding=10,
                pagination_menu_pos='auto',
                rows_num=10,
            )

            self.root.ids.card_layout.clear_widgets()
            self.root.ids.card_layout.add_widget(table)

    def on_start(self):
        file = pathlib.Path('Backend_data.xlsx')
        if not file.exists():
            file = Workbook()
            sheet = file.active
            sheet['A1'] = "Name"
            sheet['B1'] = "Standard"
            file.save('Backend_data.xlsx')

        self.reset_table()

    def reset_table(self):
        data = self.read_excel_data('Backend_data.xlsx')
        cols = ["Name", "Standard"]
        values = data

        table = MDDataTable(
            pos_hint={'center_x': 0.2, 'center_y': 0.2},
            column_data=[(col, dp(35)) for col in cols],
            row_data=values,
            use_pagination=True,
            padding=10,
            pagination_menu_pos='auto',
            rows_num=10,
        )
        self.root.ids.card_layout.clear_widgets()
        self.root.ids.card_layout.add_widget(table)

    def read_excel_data(self, filename):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        return data


if __name__ == "__main__":
    LoginApp().run()
